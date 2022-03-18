import os #フォルダを作成する
from openpyxl import load_workbook #エクセルを操作する
from tkinter import messagebox #メッセージを表示する(GUIを操作する)

#'Lecture.xlsx'を読み込み、変数wbに設定する
filepath = 'Lecture.xlsx'
wb = load_workbook(filename=filepath)
#ワークシートを読み込む
ws = wb['Sheet1']

#'B1'セルの値を読み取る
path = ws['B1'].value
#変数cmaxに最終行を設定する
cmax = ws.max_row
range = 'B4:B' + str(cmax)

#同じフォルダが存在するか確認する
existedFolder = []

for value in ws[range]:
    folderPath = path + '\\' + value[0].value
    if os.path.exists(folderPath) == False:
        os.mkdir(folderPath)
    else:
        existedFolder.append(value[0].value)

if len(existedFolder) > 0:
    #showinfo(title,message)
    #\n\nは2回改行
    #joinはリストの要素を結合する
    messagebox.showinfo('存在するフォルダがあります', 
                        '以下のフォルダは既に存在しています。　\n\n'+' '.join(existedFolder))